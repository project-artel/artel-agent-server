"""SDK 리포트 하나를 사람이 읽는 통합문서로.

    .venv/bin/python scripts/spec_workbook.py <report.json> [out.xlsx]

`/internal/specs/v2/generate` 와 같은 것을 만든다 — 같은 서비스 함수를 부르므로
API 를 띄우지 않고 본 것이 API 가 낼 것과 다를 수 없다.

시트 순서는 읽는 사람이 무엇부터 보는가로 정한다: 실행할 수 있는 것, 검토할 것,
상태가 이어지는 흐름, 그다음 화면별로 훑을 수 있게 한 장씩.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.specs_v2.service import generate_spec_payload

HEAD = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(color="FFFFFF", bold=True)
# 판정을 색으로도 말한다. 열이 스무 개라 status 한 칸을 눈으로 찾기 어렵다.
STATUS = {
    "ready": PatternFill("solid", fgColor="E2EFDA"),
    "candidate": PatternFill("solid", fgColor="FFF2CC"),
    "review": PatternFill("solid", fgColor="FCE4D6"),
    "unsupported": PatternFill("solid", fgColor="E7E6E6"),
}
# 사람이 먼저 읽는 네 열. 이 순서로 앞에 놓고 나머지는 뒤로 민다 — 어느 화면의
# 이야기인지가 먼저 오지 않으면 조건부터 읽게 된다.
LEAD = ["scene", "precondition", "test_step", "expected_result", "status"]
WIDE = {"precondition": 46, "test_step": 34, "expected_result": 46, "supporting_state": 40}


def ordered(columns: list[str]) -> list[str]:
    return [name for name in LEAD if name in columns] + [
        name for name in columns if name not in LEAD
    ]
# 화면별 시트는 판정에 필요한 것만. 전체 열은 위 두 시트에 있다.
NARROW = [
    "scene",
    "precondition",
    "test_step",
    "expected_result",
    "status",
    "ui_text",
    "ui_sprite",
    "review_reason",
    "spec_id",
    "evidence",
]


def sheet(book: Workbook, title: str, rows: list[dict], columns: list[str] | None = None):
    page = book.create_sheet(title[:31])
    if not rows:
        page["A1"] = "없음"
        return page
    columns = ordered(columns or list(rows[0].keys()))
    page.append(columns)
    for cell in page[1]:
        cell.fill, cell.font = HEAD, HEAD_FONT
    for row in rows:
        page.append([row.get(name, "") for name in columns])
    if "status" in columns:
        at = columns.index("status") + 1
        for line in range(2, len(rows) + 2):
            fill = STATUS.get(page.cell(line, at).value)
            if fill:
                page.cell(line, at).fill = fill
    for index, name in enumerate(columns, start=1):
        width = WIDE.get(name)
        if width is None:
            longest = max([len(name)] + [len(str(row.get(name, ""))) for row in rows])
            width = min(max(longest + 2, 10), 30)
        page.column_dimensions[get_column_letter(index)].width = width
    page.freeze_panes = "A2"
    page.auto_filter.ref = page.dimensions
    for line in range(2, len(rows) + 2):
        page.row_dimensions[line].height = 30
        for cell in page[line]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return page


def workbook(payload: dict) -> Workbook:
    ready = payload["ready_specs"]
    review = payload["review_specs"]
    flows = payload["connected_flows"]
    summary_rows = payload["summary"]

    book = Workbook()
    book.remove(book.active)

    summary = book.create_sheet("Summary")
    for line in [
        (f"{payload['artifact']} 독립 명세 요약", None),
        (None, None),
        ("산출물", payload["artifact"]),
        ("SDK capture", payload["capture"]),
        ("Build evidence", payload["build_evidence"]),
        (None, None),
        ("구분", "건수"),
        ("Ready specs", summary_rows["ready_specs"]),
        ("Candidate", summary_rows["candidate_specs"]),
        ("Review", summary_rows["review_specs"]),
        ("Unsupported", summary_rows["unsupported_specs"]),
        ("Connected flow rows", summary_rows["connected_flows"]),
        (None, None),
        ("이 문서는 다른 SDK 산출물을 참조하지 않습니다.", None),
    ]:
        summary.append(list(line))
    summary.column_dimensions["A"].width = 44
    summary.column_dimensions["B"].width = 20
    summary["A1"].font = Font(bold=True, size=13)
    summary["A7"].font = summary["B7"].font = Font(bold=True)

    sheet(book, "Ready Specs", ready)
    sheet(book, "Review Queue", review)
    sheet(book, "Connected Flows", flows)
    for scene in dict.fromkeys(row["scene"] for row in ready):
        sheet(book, scene, [row for row in ready if row["scene"] == scene], NARROW)
    return book


def main(report_path: str, out_path: str | None = None) -> None:
    with open(report_path, encoding="utf-8") as handle:
        payload = generate_spec_payload(json.load(handle))
    target = Path(out_path or Path(report_path).with_suffix(".xlsx").name)
    workbook(payload).save(target)
    print(f"{target}  {payload['artifact']}  {payload['summary']}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        raise SystemExit(2)
    main(*sys.argv[1:3])
